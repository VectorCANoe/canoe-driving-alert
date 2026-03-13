#!/usr/bin/env python3
"""
Markdown to Excel Converter for Extra Documents
Converts specific Markdown files for the current driving-alert-workproducts baseline.
"""

import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Reusing the parsing logic from md_to_excel.py
def convert_md_to_excel(md_file, excel_file):
    """Convert Markdown file to Excel"""
    print(f"Converting: {md_file.name} -> {excel_file.name}")

    try:
        with open(md_file, 'r', encoding='utf-8') as f:
            content = f.read()
    except FileNotFoundError:
        print(f"✗ File not found: {md_file}")
        return

    lines = content.split('\n')

    wb = Workbook()
    ws = wb.active
    ws.title = "Content"

    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, size=14, color="1F4E78")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    current_row = 1
    in_table = False

    i = 0
    while i < len(lines):
        line = lines[i].strip()

        # Title (# Header)
        if line.startswith('# '):
            ws.cell(row=current_row, column=1, value=line[2:])
            ws.cell(row=current_row, column=1).font = title_font
            current_row += 1

        # Subtitle (## Header)
        elif line.startswith('## '):
            ws.cell(row=current_row, column=1, value=line[3:])
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
            current_row += 1

        # Subheader (### Header)
        elif line.startswith('### '):
            ws.cell(row=current_row, column=1, value=line[4:])
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=11)
            current_row += 1

        # Table
        elif line.startswith('|') and not in_table:
            # Collect all table lines
            table_lines = []
            j = i
            while j < len(lines):
                tline = lines[j].strip()
                if tline.startswith('|') and tline.endswith('|'):
                    # Skip separator line
                    if not re.match(r'\|[\s\-:]+\|', tline):
                        cells = [cell.strip() for cell in tline[1:-1].split('|')]
                        table_lines.append(cells)
                    j += 1
                else:
                    break

            # Write table to Excel
            if table_lines:
                # Header row
                for col_idx, cell_value in enumerate(table_lines[0], start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=cell_value)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border

                current_row += 1

                # Data rows
                for row_data in table_lines[1:]:
                    for col_idx, cell_value in enumerate(row_data, start=1):
                        cell = ws.cell(row=current_row, column=col_idx, value=cell_value)
                        cell.border = border
                        cell.alignment = Alignment(vertical='top', wrap_text=True)
                    current_row += 1

                current_row += 1  # Add spacing after table
                i = j - 1

        # Regular text
        elif line and not line.startswith('---'):
            # Remove markdown formatting
            clean_line = re.sub(r'\*\*(.+?)\*\*', r'\1', line)  # Bold
            clean_line = re.sub(r'\*(.+?)\*', r'\1', clean_line)  # Italic
            clean_line = re.sub(r'`(.+?)`', r'\1', clean_line)  # Code

            if clean_line:
                ws.cell(row=current_row, column=1, value=clean_line)
                current_row += 1

        # Empty line
        elif not line:
            current_row += 1

        i += 1

    # Auto-adjust column widths
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column = get_column_letter(col_idx)

        for cell in ws[column]:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max_length + 2, 80)
        ws.column_dimensions[column].width = adjusted_width

    # Save
    wb.save(excel_file)
    print(f"✓ Created: {excel_file}")

def main():
    repo_root = Path(__file__).resolve().parents[2]
    source_dir = repo_root / "driving-alert-workproducts"
    output_dir = source_dir / "excel" / "excel_extra"
    output_dir.mkdir(parents=True, exist_ok=True)

    # Specific files to convert
    target_files = ["00_VModel_Mapping.md", "Schedule_calender.md"]

    print(f"\nProcessing specific files: {', '.join(target_files)}\n")

    for filename in target_files:
        md_file = source_dir / filename
        excel_file = output_dir / f"{md_file.stem}.xlsx"

        try:
            convert_md_to_excel(md_file, excel_file)
        except Exception as e:
            print(f"✗ Error converting {md_file.name}: {e}")

    print(f"\n✅ Specific conversion complete! Files saved to: {output_dir}")

if __name__ == "__main__":
    main()
