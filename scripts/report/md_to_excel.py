#!/usr/bin/env python3
"""
Markdown to Excel Converter
Converts Markdown files to Excel format, preserving tables and structure
"""

import re
import sys
import argparse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SEPARATOR_ROW_RE = re.compile(r'^\|(?:\s*:?-{3,}:?\s*\|)+\s*$')


def is_separator_row(line: str) -> bool:
    """Return True only for markdown table separator rows like |---|:---:|."""
    return bool(SEPARATOR_ROW_RE.match(line.strip()))


def parse_markdown_table(lines):
    """Parse markdown table into list of rows"""
    table_lines = []
    in_table = False

    for line in lines:
        line = line.strip()
        if line.startswith('|') and line.endswith('|'):
            # Skip separator lines (e.g., |---|---|)
            if not is_separator_row(line):
                # Remove leading/trailing pipes and split
                cells = [cell.strip() for cell in line[1:-1].split('|')]
                table_lines.append(cells)
                in_table = True
        elif in_table:
            break

    return table_lines

def write_content_to_sheet(content, ws):
    """Write markdown content to a worksheet."""
    lines = content.split('\n')

    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, size=14, color="1F4E78")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    current_row = 1
    in_table = False

    i = 0
    while i < len(lines):
        line = lines[i].strip()

        # Title (# Header)
        if line.startswith('# '):
            ws.cell(row=current_row, column=1, value=line[2:])
            ws.cell(row=current_row, column=1).font = title_font
            current_row += 1

        # Subtitle (## Header)
        elif line.startswith('## '):
            ws.cell(row=current_row, column=1, value=line[3:])
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
            current_row += 1

        # Subheader (### Header)
        elif line.startswith('### '):
            ws.cell(row=current_row, column=1, value=line[4:])
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=11)
            current_row += 1

        # Table
        elif line.startswith('|') and not in_table:
            # Collect all table lines
            table_lines = []
            j = i
            while j < len(lines):
                tline = lines[j].strip()
                if tline.startswith('|') and tline.endswith('|'):
                    # Skip separator line
                    if not is_separator_row(tline):
                        cells = [cell.strip() for cell in tline[1:-1].split('|')]
                        table_lines.append(cells)
                    j += 1
                else:
                    break

            # Write table to Excel
            if table_lines:
                # Header row
                for col_idx, cell_value in enumerate(table_lines[0], start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=cell_value)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border

                current_row += 1

                # Data rows
                for row_data in table_lines[1:]:
                    for col_idx, cell_value in enumerate(row_data, start=1):
                        cell = ws.cell(row=current_row, column=col_idx, value=cell_value)
                        cell.border = border
                        cell.alignment = Alignment(vertical='top', wrap_text=True)
                    current_row += 1

                current_row += 1  # Add spacing after table
                i = j - 1

        # Regular text
        elif line and not line.startswith('---'):
            # Remove markdown formatting
            clean_line = re.sub(r'\*\*(.+?)\*\*', r'\1', line)  # Bold
            clean_line = re.sub(r'\*(.+?)\*', r'\1', clean_line)  # Italic
            clean_line = re.sub(r'`(.+?)`', r'\1', clean_line)  # Code

            if clean_line:
                ws.cell(row=current_row, column=1, value=clean_line)
                current_row += 1

        # Empty line
        elif not line:
            current_row += 1

        i += 1

    # Auto-adjust column widths
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column = get_column_letter(col_idx)

        for cell in ws[column]:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max_length + 2, 80)
        ws.column_dimensions[column].width = adjusted_width


def sanitize_sheet_name(raw_name, used_names):
    """Return a unique worksheet name within Excel 31-char constraints."""
    # Excel forbidden chars: : \ / ? * [ ]
    cleaned = re.sub(r'[:\\\\/?*\\[\\]]', '_', raw_name).strip()
    if not cleaned:
        cleaned = "Sheet"
    base = cleaned[:31]
    name = base
    suffix = 1
    while name in used_names:
        tail = f"_{suffix}"
        name = f"{base[:31 - len(tail)]}{tail}"
        suffix += 1
    used_names.add(name)
    return name


def convert_md_to_excel(md_file, excel_file):
    """Convert one markdown file into one Excel file."""
    print(f"Converting: {md_file.name} -> {excel_file.name}")

    with open(md_file, 'r', encoding='utf-8') as f:
        content = f.read()

    wb = Workbook()
    ws = wb.active
    ws.title = "Content"
    write_content_to_sheet(content, ws)
    wb.save(excel_file)
    print(f"✓ Created: {excel_file}")


def convert_md_dir_to_single_workbook(md_files, workbook_file):
    """Convert multiple markdown files into a single workbook (one sheet per file)."""
    wb = Workbook()
    wb.remove(wb.active)
    used_names = set()

    for md_file in md_files:
        with open(md_file, 'r', encoding='utf-8') as f:
            content = f.read()
        sheet_name = sanitize_sheet_name(md_file.stem, used_names)
        ws = wb.create_sheet(title=sheet_name)
        write_content_to_sheet(content, ws)
        print(f"Added sheet: {sheet_name} ({md_file.name})")

    workbook_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook_file)
    print(f"\n✅ Single workbook created: {workbook_file}")

def main():
    parser = argparse.ArgumentParser(
        description="Convert markdown files in a directory to Excel files."
    )
    parser.add_argument(
        "source_dir",
        help="Directory containing markdown files (*.md).",
    )
    parser.add_argument(
        "output_dir",
        nargs="?",
        default=None,
        help="Output directory for xlsx files (default: <source_dir>/excel).",
    )
    parser.add_argument(
        "--single-workbook",
        dest="single_workbook",
        default=None,
        help="Create one workbook with multiple sheets (path to output xlsx).",
    )
    args = parser.parse_args()

    source_dir = Path(args.source_dir).expanduser().resolve()
    output_dir = (
        Path(args.output_dir).expanduser().resolve()
        if args.output_dir
        else source_dir / "excel"
    )

    if not source_dir.exists() or not source_dir.is_dir():
        print(f"Source directory not found or not a directory: {source_dir}")
        sys.exit(1)

    # Find all MD files
    # Default submission behavior: exclude README-style index docs from workbook sheets.
    md_files = sorted(
        p for p in source_dir.glob("*.md")
        if not p.stem.upper().startswith("README")
    )

    print(f"\nFound {len(md_files)} files to convert in {source_dir}:")
    for md_file in md_files:
        print(f"  - {md_file.name}")

    if not md_files:
        print("No markdown files found!")
        return

    if args.single_workbook:
        workbook_file = Path(args.single_workbook).expanduser().resolve()
        print("\nConverting to single workbook...\n")
        convert_md_dir_to_single_workbook(md_files, workbook_file)
        return

    output_dir.mkdir(parents=True, exist_ok=True)
    print("\nConverting...\n")
    for md_file in md_files:
        excel_file = output_dir / f"{md_file.stem}.xlsx"
        try:
            convert_md_to_excel(md_file, excel_file)
        except Exception as e:
            print(f"✗ Error converting {md_file.name}: {e}")

    print(f"\n✅ Conversion complete! Files saved to: {output_dir}")

if __name__ == "__main__":
    main()
