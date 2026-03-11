#!/usr/bin/env python3
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
FINAL_DOCS = ROOT / "final-docs"
OUTPUT = ROOT / "excel" / "submission_final_all_in_one.xlsx"

ORDER = [
    "00d_HARA_Worksheet.md",
    "00e_ECU_Naming_Standard.md",
    "00f_CAN_ID_Allocation_Standard.md",
    "01_Requirements.md",
    "02_Concept_design.md",
    "03_Function_definition.md",
    "0301_SysFuncAnalysis.md",
    "0302_NWflowDef.md",
    "0303_Communication_Specification.md",
    "0304_System_Variables.md",
    "04_SW_Implementation.md",
    "05_Unit_Test.md",
    "06_Integration_Test.md",
    "07_System_Test.md",
]


def clean_text(text: str) -> str:
    text = text.rstrip()
    text = re.sub(r"^#{1,6}\s*", "", text)
    text = re.sub(r"^\>\s*", "", text)
    text = re.sub(r"^-\s*", "", text)
    text = re.sub(r"^\d+\.\s*", "", text)
    text = re.sub(r"\*\*(.*?)\*\*", r"\1", text)
    text = re.sub(r"`([^`]*)`", r"\1", text)
    text = re.sub(r"!\[(.*?)\]\((.*?)\)", r"Image: \2", text)
    return text.strip()


def is_table_line(line: str) -> bool:
    stripped = line.strip()
    return stripped.startswith("|") and stripped.endswith("|")


def is_separator_line(line: str) -> bool:
    stripped = line.strip()
    if not (stripped.startswith("|") and stripped.endswith("|")):
        return False
    cells = [cell.strip() for cell in stripped.strip("|").split("|")]
    return all(c and set(c) <= {":", "-"} for c in cells)


def parse_table_row(line: str) -> list[str]:
    return [clean_text(cell) for cell in line.strip().strip("|").split("|")]


def unique_sheet_title(used: set[str], stem: str) -> str:
    title = stem[:31]
    if title not in used:
        used.add(title)
        return title

    idx = 2
    while True:
        suffix = f"_{idx}"
        candidate = f"{stem[:31 - len(suffix)]}{suffix}"
        if candidate not in used:
            used.add(candidate)
            return candidate
        idx += 1


def write_lines(ws, lines: list[str]) -> None:
    row = 1
    i = 0
    while i < len(lines):
        line = lines[i].rstrip("\n")
        stripped = line.strip()

        if not stripped:
            row += 1
            i += 1
            continue

        if is_table_line(stripped):
            header_written = False
            while i < len(lines) and is_table_line(lines[i].strip()):
                current = lines[i].strip()
                if is_separator_line(current):
                    i += 1
                    continue
                cells = parse_table_row(current)
                for col, cell in enumerate(cells, start=1):
                    ws.cell(row=row, column=col, value=cell)
                if not header_written:
                    for col in range(1, len(cells) + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill("solid", fgColor="D9EAD3")
                    header_written = True
                row += 1
                i += 1
            row += 1
            continue

        text = clean_text(line)
        if text:
            ws.cell(row=row, column=1, value=text)
            if line.lstrip().startswith("#"):
                ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        i += 1

    ws.freeze_panes = "A1"
    for column_cells in ws.columns:
        max_len = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=0)
        width = min(max(max_len + 2, 12), 48)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    used_titles: set[str] = set()
    for name in ORDER:
        path = FINAL_DOCS / name
        if not path.exists():
            continue
        ws = wb.create_sheet(unique_sheet_title(used_titles, path.stem))
        write_lines(ws, path.read_text(encoding="utf-8").splitlines())

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)


if __name__ == "__main__":
    build()
